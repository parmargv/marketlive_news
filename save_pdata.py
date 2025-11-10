import os
import pandas as pd
import openpyxl

def save_pro(FII_DATA,PRO_DATA,DII_DATA,CLI_DATA,NET_DATA):
    absolute_path = os.path.dirname(__file__)
    file_path = os.path.join(absolute_path, 'p_data.xlsx')
    wb = openpyxl.load_workbook(file_path)
    sh1 = wb['FII']
    sh2= wb['PRO']
    sh3 =wb['NET']
    sh4 =wb['DII']
    sh5 =wb['CLI']


    headers = ['DATE', 'FII_CALL','FII_PUT','FII_NET']
    for col_num, header in enumerate(headers, start=1):
        sh1.cell(row=1, column=col_num, value=header)
    # Find the next available row starting from row 2
    next_row = 2
    while sh1.cell(row=next_row, column=1).value is not None:
        next_row += 1
    # Write data to the next available row
    for col_num, value in enumerate(FII_DATA, start=1):
        sh1.cell(row=next_row, column=col_num, value=value)
    # Save the workbook
    wb.save('p_data.xlsx')

    headers = ['Date', 'PRO_CALL', 'PRO_PUT', 'PRO_NET']
    for col_num, header in enumerate(headers, start=1):
        sh2.cell(row=1, column=col_num, value=header)
    # Find the next available row starting from row 2
    next_row = 2
    while sh2.cell(row=next_row, column=1).value is not None:
        next_row += 1
    # Write data to the next available row
    for col_num, value in enumerate(PRO_DATA, start=1):
        sh2.cell(row=next_row, column=col_num, value=value)
    # Save the workbook
    wb.save('p_data.xlsx')

    headers = ['DATE','NET']
    for col_num, header in enumerate(headers, start=1):
        sh3.cell(row=1, column=col_num, value=header)
    # Find the next available row starting from row 2
    next_row = 2
    while sh3.cell(row=next_row, column=1).value is not None:
        next_row += 1
    # Write data to the next available row
    for col_num, value in enumerate(NET_DATA, start=1):
        sh3.cell(row=next_row, column=col_num, value=value)
    # Save the workbook
    wb.save('p_data.xlsx')
    wb.close()

    headers = ['Date', 'DII_CALL', 'DII_PUT', 'DII_NET']
    for col_num, header in enumerate(headers, start=1):
        sh4.cell(row=1, column=col_num, value=header)
    # Find the next available row starting from row 2
    next_row = 2
    while sh4.cell(row=next_row, column=1).value is not None:
        next_row += 1
    # Write data to the next available row
    for col_num, value in enumerate(DII_DATA, start=1):
        sh4.cell(row=next_row, column=col_num, value=value)
    # Save the workbook
    wb.save('p_data.xlsx')

    headers = ['Date', 'CLI_CALL', 'CLI_PUT', 'CLI_NET']
    for col_num, header in enumerate(headers, start=1):
        sh5.cell(row=1, column=col_num, value=header)
    # Find the next available row starting from row 2
    next_row = 2
    while sh5.cell(row=next_row, column=1).value is not None:
        next_row += 1
    # Write data to the next available row
    for col_num, value in enumerate(CLI_DATA, start=1):
        sh5.cell(row=next_row, column=col_num, value=value)
    # Save the workbook
    wb.save('p_data.xlsx')

def read_data():
    absolute_path = os.path.dirname(__file__)
    file_path = os.path.join(absolute_path, 'p_data.xlsx')
    wb = openpyxl.load_workbook(file_path)
    sh1 = wb['FII']
    sh2 = wb['PRO']
    sh3 = wb['NET']
    sh4 = wb['DII']
    sh5 = wb['CLI']
    data1 = []
    data2 =[]
    data3 = []
    data4=[]
    data5=[]

    for row in sh1.iter_rows(min_row=2,max_col=4, values_only=True):
        if not all(cell is None for cell in row):
            data1.append(row)
    df1 = pd.DataFrame(data1, columns=['DATE', 'FII_CALL', 'FII_PUT', 'FII_NET'])
    df1= df1.sort_values(by='DATE')
    df1.reset_index(drop=True, inplace=True)

    for row in sh2.iter_rows(min_row=2,max_col=4, values_only=True):
        if not all(cell is None for cell in row):
            data2.append(row)
    df2 = pd.DataFrame(data2, columns=['DATE', 'PRO_CALL', 'PRO_PUT', 'PRO_NET'])
    df2 = df2.sort_values(by='DATE')
    df2.reset_index(drop=True, inplace=True)

    for row in sh3.iter_rows(min_row=2,max_col=2, values_only=True):
        if not all(cell is None for cell in row):
            data3.append(row)
    df3 = pd.DataFrame(data3, columns=['DATE','NET'])
    df3 = df3.sort_values(by='DATE')
    df3.reset_index(drop=True, inplace=True)

    for row in sh4.iter_rows(min_row=2,max_col=4, values_only=True):
        if not all(cell is None for cell in row):
            data4.append(row)
    df4 = pd.DataFrame(data4, columns=['DATE', 'DII_CALL', 'DII_PUT', 'DII_NET'])
    df4 = df4.sort_values(by='DATE')
    df4.reset_index(drop=True, inplace=True)

    for row in sh5.iter_rows(min_row=2,max_col=4, values_only=True):
        if not all(cell is None for cell in row):
            data5.append(row)
    df5 = pd.DataFrame(data5, columns=['DATE', 'CLI_CALL', 'CLI_PUT', 'CLI_NET'])
    df5 = df5.sort_values(by='DATE')
    df5.reset_index(drop=True, inplace=True)

    return df1,df2,df3,df4,df5
def clear_data():
    absolute_path = os.path.dirname(__file__)
    file_path = os.path.join(absolute_path, 'p_data.xlsx')
    wb = openpyxl.load_workbook(file_path)
    sh1 = wb['FII']
    sh2 = wb['PRO']
    sh3 = wb['NET']
    sh4 = wb['DII']
    sh5 = wb['CLI']

    start_row = 2
    end_row = 20
    start_col = 1
    end_col = 4

    # Clear the range by setting each cell's value to None
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sh1.cell(row=row, column=col)
            cell.value = None

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sh2.cell(row=row, column=col)
            cell.value = None

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sh3.cell(row=row, column=col)
            cell.value = None

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sh4.cell(row=row, column=col)
            cell.value = None

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sh5.cell(row=row, column=col)
            cell.value = None

    # Save the workbook
    wb.save('p_data.xlsx')

    wb.close()